import os
import sys
import json
import random
import time
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QCalendarWidget, QPushButton, QLineEdit, QTableWidget, QHeaderView, 
                             QSpinBox, QMessageBox, QTableWidgetItem, QCheckBox, QLabel, 
                             QGridLayout, QFileDialog, QProgressDialog, QInputDialog)
from PyQt5.QtCore import Qt, QDate, QTimer, pyqtSignal, QObject
from PyQt5.QtGui import QColor
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def get_data_dir():
    if getattr(sys, 'frozen', False):
        # We are running in a bundle
        bundle_dir = sys._MEIPASS
    else:
        # We are running in a normal Python environment
        bundle_dir = os.path.dirname(os.path.abspath(__file__))
    return bundle_dir

def get_user_data_dir():
    if sys.platform == 'win32':
        return os.path.join(os.environ['APPDATA'], 'SchedulingApp')
    elif sys.platform == 'darwin':
        return os.path.expanduser('~/Library/Application Support/SchedulingApp')
    else:
        return os.path.expanduser('~/.schedulingapp')

class SchedulerSignals(QObject):
    progress = pyqtSignal(int, int, float)
    finished = pyqtSignal(np.ndarray, float)

class OptimizedGeneticScheduler:
    def __init__(self, employees, days, daily_demand, rest_days, start_date, manual_schedule, population_size=1000, mutation_rate=0.01):
        self.employees = employees
        self.days = days
        self.daily_demand = daily_demand
        self.rest_days = rest_days
        self.start_date = start_date
        self.manual_schedule = manual_schedule
        self.population_size = population_size
        self.mutation_rate = mutation_rate
        self.current_generation = 0
        self.best_fitness = float('-inf')
        self.signals = SchedulerSignals()

    def generate_individual(self):
        schedule = np.zeros((self.employees, self.days), dtype=int)
        for emp in range(self.employees):
            work_days = self.days - self.rest_days[emp]
            available_days = [day for day in range(self.days) if self.manual_schedule[emp][day] == -1]
            schedule[emp, random.sample(available_days, min(work_days, len(available_days)))] = 1
        
        # Apply manual schedule
        for emp in range(self.employees):
            for day in range(self.days):
                if self.manual_schedule[emp][day] != -1:
                    schedule[emp, day] = self.manual_schedule[emp][day]
        
        return schedule

    def fitness(self, individual):
        score = 0
        
        # Check daily demand
        for day in range(self.days):
            diff = abs(np.sum(individual[:, day]) - self.daily_demand[day])
            score -= diff * 10  # Heavily penalize not meeting daily demand

        # Check rest days and consecutive work days
        for emp in range(self.employees):
            rest_days = np.sum(individual[emp] == 0)
            if rest_days < self.rest_days[emp]:
                score -= (self.rest_days[emp] - rest_days) * 5

            consecutive_work = 0
            consecutive_rest = 0
            for day in range(self.days):
                if individual[emp, day] == 1:
                    consecutive_work += 1
                    consecutive_rest = 0
                    if consecutive_work > 6:
                        score -= 5  # Penalize more than 6 consecutive work days
                else:
                    consecutive_work = 0
                    consecutive_rest += 1
                    if consecutive_rest > 2:
                        score -= 3 * (consecutive_rest - 2)  # Penalize more than 2 consecutive rest days

        # Prioritize weekend rest
        current_date = self.start_date
        for day in range(self.days):
            if current_date.weekday() >= 5:  # Saturday or Sunday
                score += np.sum(individual[:, day] == 0) * 2
            current_date += timedelta(days=1)

        # Encourage consecutive rest days (but not more than 2)
        for emp in range(self.employees):
            for day in range(self.days - 1):
                if individual[emp, day] == 0 and individual[emp, day+1] == 0:
                    score += 1
                    if day < self.days - 2 and individual[emp, day+2] == 1:
                        score += 1  # Extra bonus for exactly 2 consecutive rest days

        return score

    def tournament_selection(self, population, k=3):
        selected = random.sample(population, k)
        return max(selected, key=self.fitness)

    def crossover(self, parent1, parent2):
        child = parent1.copy()
        for emp in range(self.employees):
            if random.random() < 0.5:
                child[emp, :] = parent2[emp, :]
        
        # Ensure manual schedule is maintained
        for emp in range(self.employees):
            for day in range(self.days):
                if self.manual_schedule[emp][day] != -1:
                    child[emp, day] = self.manual_schedule[emp][day]
        
        return child

    def mutate(self, individual):
        for emp in range(self.employees):
            if random.random() < self.mutation_rate:
                available_days = [day for day in range(self.days) if self.manual_schedule[emp][day] == -1]
                if len(available_days) >= 2:
                    day1, day2 = random.sample(available_days, 2)
                    individual[emp, day1], individual[emp, day2] = individual[emp, day2], individual[emp, day1]
        return individual

    def evolve(self, generations=1000, patience=50, time_limit=60):
        population = [self.generate_individual() for _ in range(self.population_size)]
        self.best_fitness = float('-inf')
        best_solution = None
        generations_without_improvement = 0
        start_time = time.time()

        for gen in range(generations):
            self.current_generation = gen + 1
            population = sorted(population, key=self.fitness, reverse=True)
            current_best_fitness = self.fitness(population[0])

            if current_best_fitness > self.best_fitness:
                self.best_fitness = current_best_fitness
                best_solution = population[0].copy()
                generations_without_improvement = 0
            else:
                generations_without_improvement += 1

            # Emit progress signal
            self.signals.progress.emit(self.current_generation, generations, self.best_fitness)

            if generations_without_improvement >= patience or time.time() - start_time > time_limit:
                break

            # Adaptive population size
            elapsed_time = time.time() - start_time
            if elapsed_time > 5 and self.current_generation < 10:
                self.population_size = max(100, self.population_size // 2)
                population = population[:self.population_size]

            new_population = population[:2]  # Elitism

            while len(new_population) < self.population_size:
                parent1 = self.tournament_selection(population)
                parent2 = self.tournament_selection(population)
                child = self.crossover(parent1, parent2)
                child = self.mutate(child)
                new_population.append(child)

            population = new_population

            # Adaptive mutation rate
            self.mutation_rate = max(0.001, min(0.1, self.mutation_rate * (1 + (generations_without_improvement - 10) * 0.01)))
        
        self.signals.finished.emit(best_solution, self.best_fitness)
        return best_solution, self.best_fitness

class SchedulingApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("智能排班系统")
        self.setGeometry(100, 100, 1500, 900)
        # Ensure user data directory exists
        self.user_data_dir = get_user_data_dir()
        os.makedirs(self.user_data_dir, exist_ok=True)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 日历选择
        calendar_layout = QHBoxLayout()
        self.start_calendar = QCalendarWidget()
        self.end_calendar = QCalendarWidget()
        calendar_layout.addWidget(self.start_calendar)
        calendar_layout.addWidget(self.end_calendar)
        main_layout.addLayout(calendar_layout)

        # 员工管理区域
        employee_management_layout = QVBoxLayout()
        input_layout = QHBoxLayout()
        self.employee_input = QLineEdit()
        self.add_employee_button = QPushButton("+")
        input_layout.addWidget(self.employee_input)
        input_layout.addWidget(self.add_employee_button)
        employee_management_layout.addLayout(input_layout)

        self.employee_grid = QGridLayout()
        employee_management_layout.addLayout(self.employee_grid)
        main_layout.addLayout(employee_management_layout)

        # 排班表格
        self.schedule_table = QTableWidget()
        self.schedule_table.setMinimumHeight(500)
        main_layout.addWidget(self.schedule_table)

        # 按钮布局
        button_layout = QHBoxLayout()
        self.schedule_button = QPushButton("智能排班")
        self.reset_button = QPushButton("重置排班")
        self.export_button = QPushButton("导出排班表")
        button_layout.addWidget(self.schedule_button)
        button_layout.addWidget(self.reset_button)
        button_layout.addWidget(self.export_button)
        main_layout.addLayout(button_layout)

        self.all_employees = self.load_employees()
        self.selected_employees = []
        self.setup_connections()
        self.update_employee_grid()

        self.progress_dialog = None
        self.scheduler = None
        self.is_scheduling = False
        self.signals_connected = False

    def setup_connections(self):
        self.start_calendar.selectionChanged.connect(self.update_table)
        self.end_calendar.selectionChanged.connect(self.update_table)
        self.add_employee_button.clicked.connect(self.add_employee)
        self.employee_input.returnPressed.connect(self.add_employee)
        self.schedule_button.clicked.connect(self.run_scheduling)
        self.reset_button.clicked.connect(self.reset_schedule)
        self.export_button.clicked.connect(self.export_schedule)
        self.schedule_table.cellClicked.connect(self.toggle_cell_state)

    def load_employees(self):
        json_path = os.path.join(self.user_data_dir, 'employees.json')
        if not os.path.exists(json_path):
            # If the file doesn't exist in the user data directory, copy it from the bundle
            bundle_json_path = os.path.join(get_data_dir(), 'employees.json')
            if os.path.exists(bundle_json_path):
                shutil.copy2(bundle_json_path, json_path)
            else:
                # If no file exists, create an empty one
                with open(json_path, 'w') as f:
                    json.dump([], f)

        try:
            with open(json_path, 'r') as f:
                return json.load(f)
        except json.JSONDecodeError:
            # If the file is corrupted, return an empty list
            return []

    def save_employees(self):
        json_path = os.path.join(self.user_data_dir, 'employees.json')
        with open(json_path, 'w') as f:
            json.dump(self.all_employees, f)

    def update_employee_grid(self):
        for i in reversed(range(self.employee_grid.count())): 
            self.employee_grid.itemAt(i).widget().setParent(None)
        
        row = 0
        col = 0
        for employee in self.all_employees:
            checkbox = QCheckBox(employee)
            checkbox.stateChanged.connect(self.update_selected_employees)
            self.employee_grid.addWidget(checkbox, row, col)
            col += 1
            if col == 5:
                col = 0
                row += 1

    def update_selected_employees(self):
        self.selected_employees = [checkbox.text() for checkbox in self.findChildren(QCheckBox) if checkbox.isChecked()]
        self.update_table()

    def add_employee(self):
        employee_name = self.employee_input.text().strip()
        if employee_name and employee_name not in self.all_employees:
            self.all_employees.append(employee_name)
            self.employee_input.clear()
            self.save_employees()
            self.update_employee_grid()
        elif employee_name in self.all_employees:
            QMessageBox.warning(self, "重复员工", "该员工已存在")
        else:
            QMessageBox.warning(self, "无效输入", "请输入有效的员工姓名")

    def update_table(self):
        start_date = self.start_calendar.selectedDate().toPyDate()
        end_date = self.end_calendar.selectedDate().toPyDate()
        
        if start_date > end_date:
            QMessageBox.warning(self, "日期错误", "开始日期不能晚于结束日期")
            return

        days = (end_date - start_date).days + 1
        self.schedule_table.setColumnCount(days + 2)  # 员工名称列和休息天数列
        self.schedule_table.setRowCount(len(self.selected_employees) + 2)  # 包括标题行和每日到岗人数行

        self.schedule_table.setHorizontalHeaderItem(0, QTableWidgetItem("员工"))
        self.schedule_table.setHorizontalHeaderItem(1, QTableWidgetItem("休息天数"))
        current_date = start_date
        for i in range(days):
            header_item = QTableWidgetItem(f"{current_date.strftime('%m-%d')}\n{['一','二','三','四','五','六','日'][current_date.weekday()]}")
            self.schedule_table.setHorizontalHeaderItem(i + 2, header_item)
            current_date += timedelta(days=1)

        self.schedule_table.setVerticalHeaderItem(0, QTableWidgetItem("每日需求"))
        for i, employee in enumerate(self.selected_employees):
            self.schedule_table.setVerticalHeaderItem(i + 1, QTableWidgetItem(str(i + 1)))
            self.schedule_table.setItem(i + 1, 0, QTableWidgetItem(employee))

            rest_days_spinbox = QSpinBox()
            rest_days_spinbox.setRange(0, days)
            rest_days_spinbox.setValue(days // 7 * 2)  # Default to 2 days off per week
            self.schedule_table.setCellWidget(i + 1, 1, rest_days_spinbox)

        self.schedule_table.setVerticalHeaderItem(len(self.selected_employees) + 1, QTableWidgetItem("每日到岗人数"))

        for col in range(2, self.schedule_table.columnCount()):
            daily_demand_spinbox = QSpinBox()
            daily_demand_spinbox.setRange(0, len(self.selected_employees))
            daily_demand_spinbox.setValue(len(self.selected_employees) // 2)  # Default to half of employees
            self.schedule_table.setCellWidget(0, col, daily_demand_spinbox)

        for row in range(1, self.schedule_table.rowCount() - 1):
            for col in range(2, self.schedule_table.columnCount()):
                item = QTableWidgetItem()
                item.setBackground(Qt.white)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 使单元格不可编辑
                self.schedule_table.setItem(row, col, item)

        self.schedule_table.resizeColumnsToContents()
        self.schedule_table.resizeRowsToContents()

        self.update_statistics()  # 初始化统计数据

    def run_scheduling(self):
        if self.is_scheduling:
            return

        start_date = self.start_calendar.selectedDate().toPyDate()
        days = self.schedule_table.columnCount() - 2
        employees = len(self.selected_employees)

        if employees == 0:
            QMessageBox.warning(self, "错误", "请选择至少一名员工")
            return

        population_size, ok = QInputDialog.getInt(self, "设置种群大小", "请输入种群大小 (100-1000):", 100, 100, 1000)
        if not ok:
            return

        rest_days = [self.schedule_table.cellWidget(row, 1).value() for row in range(1, employees + 1)]
        daily_demand = [self.schedule_table.cellWidget(0, col).value() for col in range(2, self.schedule_table.columnCount())]

        manual_schedule = []
        for row in range(1, employees + 1):
            employee_schedule = []
            for col in range(2, self.schedule_table.columnCount()):
                item = self.schedule_table.item(row, col)
                if item.text() == "班":
                    employee_schedule.append(1)
                elif item.text() == "休":
                    employee_schedule.append(0)
                else:
                    employee_schedule.append(-1)
            manual_schedule.append(employee_schedule)

        self.scheduler = OptimizedGeneticScheduler(employees, days, daily_demand, rest_days, start_date, manual_schedule, population_size=population_size)
        
        self.progress_dialog = QProgressDialog("正在进行智能排班...", "取消", 0, 100, self)
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setWindowTitle("排班进度")
        self.progress_dialog.canceled.connect(self.cancel_scheduling)
        self.progress_dialog.setAutoClose(False)
        self.progress_dialog.setAutoReset(False)
        self.progress_dialog.show()

        self.current_run = 0
        self.best_overall = None
        self.best_fitness = float('-inf')
        self.total_runs = 3

        self.is_scheduling = True
        self.signals_connected = False
        self.run_next_scheduling()

    def run_next_scheduling(self):
        if self.current_run < self.total_runs and not self.progress_dialog.wasCanceled():
            self.current_run += 1
            self.connect_scheduler_signals()
            
            # Run the scheduler in a separate thread
            QTimer.singleShot(0, self.scheduler.evolve)
        else:
            self.finish_scheduling()

    def update_scheduling_progress(self, current_generation, total_generations, best_fitness):
        if self.progress_dialog is None or self.progress_dialog.wasCanceled():
            return
        progress = ((self.current_run - 1) * 100 + (current_generation / total_generations) * 100) / self.total_runs
        self.progress_dialog.setValue(int(progress))
        self.progress_dialog.setLabelText(f"正在进行第 {self.current_run} 次排班\n"
                                          f"当前代数: {current_generation}/{total_generations}\n"
                                          f"当前最佳适应度: {best_fitness:.2f}")
    
    def connect_scheduler_signals(self):
        if not self.signals_connected:
            self.scheduler.signals.progress.connect(self.update_scheduling_progress)
            self.scheduler.signals.finished.connect(self.on_scheduling_finished)
            self.signals_connected = True
    
    def disconnect_scheduler_signals(self):
        if self.signals_connected:
            try:
                self.scheduler.signals.progress.disconnect(self.update_scheduling_progress)
                self.scheduler.signals.finished.disconnect(self.on_scheduling_finished)
            except TypeError:
                # If disconnection fails, it means the signal was not connected
                pass
            self.signals_connected = False
    
    
        
    def cancel_scheduling(self):
        self.is_scheduling = False
        self.disconnect_scheduler_signals()
        if self.progress_dialog:
            self.progress_dialog.close()
        QMessageBox.information(self, "排班取消", "排班过程已被取消")

    def finish_scheduling(self):
        self.is_scheduling = False
        if self.progress_dialog:
            self.progress_dialog.close()
        
        if self.best_overall is not None:
            self.update_table_with_schedule(self.best_overall)
            self.update_statistics()
            QMessageBox.information(self, "排班完成", f"已完成排班，最佳适应度分数：{self.best_fitness:.2f}")
        else:
            QMessageBox.information(self, "排班取消", "排班过程已被取消")

    def on_scheduling_finished(self, schedule, fitness):
        if fitness > self.best_fitness:
            self.best_fitness = fitness
            self.best_overall = schedule

        self.disconnect_scheduler_signals()
        self.run_next_scheduling()

    def update_table_with_schedule(self, schedule):
        employees, days = schedule.shape
        for row in range(employees):
            for col in range(days):
                item = self.schedule_table.item(row + 1, col + 2)
                if schedule[row, col] == 1:
                    item.setBackground(QColor(144, 238, 144))  # Light green
                    item.setText("班")
                else:
                    item.setBackground(QColor(255, 182, 193))  # Light pink
                    item.setText("休")

    def update_statistics(self):
        employees = self.schedule_table.rowCount() - 2  # 减去标题行和统计行
        days = self.schedule_table.columnCount() - 2  # 减去员工名和休息天数列

        # 更新每日到岗人数
        for col in range(2, self.schedule_table.columnCount()):
            daily_count = sum(1 for row in range(1, employees + 1) if self.schedule_table.item(row, col).text() == "班")
            stat_item = self.schedule_table.item(employees + 1, col)
            if stat_item is None:
                stat_item = QTableWidgetItem()
                stat_item.setFlags(stat_item.flags() & ~Qt.ItemIsEditable)
                self.schedule_table.setItem(employees + 1, col, stat_item)
            stat_item.setText(str(daily_count))

    def toggle_cell_state(self, row, col):
        if row == 0 or col < 2:
            return
        
        item = self.schedule_table.item(row, col)
        if item.background() == Qt.white:
            item.setBackground(QColor(144, 238, 144))  # Light green
            item.setText("班")
        elif item.text() == "班":
            item.setBackground(QColor(255, 182, 193))  # Light pink
            item.setText("休")
        else:
            item.setBackground(Qt.white)
            item.setText("")

        self.update_statistics()  # 每次单元格状态改变时更新统计数据

    def reset_schedule(self):
        for row in range(1, self.schedule_table.rowCount() - 1):
            for col in range(2, self.schedule_table.columnCount()):
                item = self.schedule_table.item(row, col)
                item.setBackground(Qt.white)
                item.setText("")
        self.update_statistics()

    def export_schedule(self):
        start_date = self.start_calendar.selectedDate().toPyDate()
        end_date = self.end_calendar.selectedDate().toPyDate()
        file_name = f"排班表_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
        
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "保存排班表", file_name, "Excel Files (*.xlsx)", options=options)
        
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.title = "排班表"

            # Write headers
            ws.cell(row=1, column=1, value="员工")
            ws.cell(row=1, column=2, value="休息天数")
            for col in range(2, self.schedule_table.columnCount()):
                header_item = self.schedule_table.horizontalHeaderItem(col)
                ws.cell(row=1, column=col+1, value=header_item.text())

            # Write data
            for row in range(self.schedule_table.rowCount()):
                for col in range(self.schedule_table.columnCount()):
                    cell_value = ""
                    cell_color = None

                    if col == 0:
                        item = self.schedule_table.verticalHeaderItem(row)
                        cell_value = item.text() if item else ""
                    elif col == 1 and row > 0:
                        spinbox = self.schedule_table.cellWidget(row, col)
                        cell_value = spinbox.value() if spinbox else ""
                    else:
                        item = self.schedule_table.item(row, col)
                        if item:
                            cell_value = item.text()
                            if item.background() == QColor(144, 238, 144):  # Light green
                                cell_color = "90EE90"
                            elif item.background() == QColor(255, 182, 193):  # Light pink
                                cell_color = "FFB6C1"

                    excel_cell = ws.cell(row=row+2, column=col+1, value=cell_value)
                    if cell_color:
                        excel_cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")

            # Adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"排班表已成功导出到：\n{file_path}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SchedulingApp()
    window.show()
    sys.exit(app.exec_())